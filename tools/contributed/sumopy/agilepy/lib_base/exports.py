# Eclipse SUMO, Simulation of Urban MObility; see https://eclipse.org/sumo
# Copyright (C) 2016-2023 German Aerospace Center (DLR) and others.
# SUMOPy module
# Copyright (C) 2012-2021 University of Bologna - DICAM
# This program and the accompanying materials are made available under the
# terms of the Eclipse Public License 2.0 which is available at
# https://www.eclipse.org/legal/epl-2.0/
# This Source Code may also be made available under the following Secondary
# Licenses when the conditions for such availability set forth in the Eclipse
# Public License 2.0 are satisfied: GNU General Public License, version 2
# or later which is available at
# https://www.gnu.org/licenses/old-licenses/gpl-2.0-standalone.html
# SPDX-License-Identifier: EPL-2.0 OR GPL-2.0-or-later

# @file    exports.py
# @author  Joerg Schweizer
# @date    2012


import os
import sys
import types
import numpy as np
import agilepy.lib_base.classman as cm
import agilepy.lib_base.arrayman as am


from agilepy.lib_base.processes import Process

#global IS_EXCEL

try:
    from openpyxl import Workbook

    IS_EXCEL = True


except:
    print 'WARNING: No Exel export possible. Install openpyxl python package.'
    IS_EXCEL = False

# if 'Workbook' in dir():
#    print 'detected Exel'
#    IS_EXCEL = True
# else:
#    IS_EXCEL = False

print 'IS_EXCEL', IS_EXCEL


def export_excel(filepath, obj, ids=None, attrconfigs=None,  groupnames=None,
                 is_header=True, is_ident=False, is_timestamp=True,
                 show_parentesis=True, name_id='ID', is_export_not_save=True):

    print 'export_excel'  # ,attrconfigs,'groupnames',groupnames,

    wb = Workbook()
    if (wb.worksheets) >= 1:
        ws = wb.worksheets[0]
        ws.title = "Scalar attributes"
    else:
        ws = wb.create_sheet(title="Scalar attributes")

    attrsman = obj.get_attrsman()
    if attrconfigs is None:
        attrconfigs = attrsman.get_configs(is_all=False,  filtergroupnames=groupnames)
        #attrconfigs_cols = attrsman.get_configs(is_all = False, structs = cm.STRUCTS_COL, filtergroupnames = groupnames)
    ind_row = 1

    for attrconf in attrconfigs:
        # print '  attrconfig', attrconf.attrname,attrconf.struct
        if (attrconf.is_save() | is_export_not_save) & (attrconf.struct in cm.STRUCTS_SCALAR):
            # print '    attrconf.attrname',attrconf.attrname,'val',attrconf.get_value()
            cell = ws.cell(row=ind_row, column=1)
            cell.value = attrconf.attrname+attrconf.format_unit(show_parentesis)
            cell = ws.cell(row=ind_row, column=2)
            value = attrconf.get_value()
            tv = type(value)
            if tv in cm.NODATATYPES:
                value = str(value)
            elif tv in (types.ListType, types.TupleType, np.ndarray):
                value = str(value)

            cell.value = value
            #d.value = 3.14
            # print "      cell",ind_row,'value',cell.value
            ind_row += 1

    # check if attributes are all column attribute and indicted for save

    ws = wb.create_sheet(title="Table attributes")
    if hasattr(obj, 'get_ids'):
        if ids is None:
            ids = obj.get_ids()

        attrconfigs_checked = []
        for attrconf in attrconfigs:
            if (attrconf.is_save() | is_export_not_save) & (attrconf.struct in cm.STRUCTS_COL):
                attrconfigs_checked.append(attrconf)

        # first table row
        ind_row = 1
        ind_col = 1
        cell = ws.cell(row=ind_row, column=ind_col)
        cell.value = name_id

        for attrconf in attrconfigs_checked:

            cell = ws.cell(row=ind_row, column=ind_col)
            cell.value = attrconf.format_symbol(show_parentesis=show_parentesis)
            ind_col += 1

        # rest
        for _id in ids:
            ind_row += 1
            ind_col = 1
            cell = ws.cell(row=ind_row, column=ind_col)
            cell.value = _id

            for attrconf in attrconfigs_checked:

                cell = ws.cell(row=ind_row, column=ind_col)
                value = attrconf[_id]
                # print '  attrconfig', attrconf.attrname,type(value)
                tv = type(value)
                mt = attrconf.metatype
                if tv in cm.NODATATYPES:
                    value = str(value)

                elif mt == 'id':
                    value = attrconf.get_linktab().format_ids([value])
                elif tv in (types.ListType, types.TupleType, np.ndarray):
                    value = str(value)
                cell.value = value
                ind_col += 1

    wb.save(filepath)


class CsvExporter(Process):
    def __init__(self,  obj, ident='csvexporter', name='CSV exporter',
                 info='Export data from a CSV file into object',
                 logger=None, **kwargs):
        print 'CsvExporter.__init__'
        self._init_common(ident,
                          parent=obj,
                          name=name,
                          logger=logger,
                          info=info,
                          )

        attrsman = self.set_attrsman(cm.Attrsman(self))
        self.csvfilepath = attrsman.add(cm.AttrConf('csvfilepath', kwargs.get('csvfilepath', ''),
                                                    groupnames=['options'],
                                                    perm='rw',
                                                    name='CSV file',
                                                    wildcards='CSV file (*.csv)|*.csv|*.CSV',
                                                    metatype='filepath',
                                                    info="CSV plain text file path.",
                                                    ))

        self.sep = attrsman.add(cm.AttrConf('sep', kwargs.get('sep', ","),
                                            groupnames=['options', ],
                                            perm='rw',
                                            name='Separator',
                                            info="""Seperator used in SCV file. Exampe: ; , <space key>""",
                                            ))

        self.is_header = attrsman.add(cm.AttrConf('is_header', kwargs.get('is_header', True),
                                                  groupnames=['options', ],
                                                  perm='rw',
                                                  name='Make header',
                                                  info="""Make header with date and time.""",
                                                  ))

        self.show_parentesis = attrsman.add(cm.AttrConf('show_parentesis', kwargs.get('show_parentesis', True),
                                                        groupnames=['options', ],
                                                        perm='rw',
                                                        name='Show units in parenthesis',
                                                        info="""Show units (if any) in parenthesis.""",
                                                        ))
        # self.name_id = attrsman.add(cm.AttrConf('name_id', kwargs.get('name_id',"ID"),
        #                    groupnames = ['options',],
        #                    perm='rw',
        #                    name = 'ID string',
        #                    info = """String to be used to indicate IDs""",
        #                    ))

    def do(self):
        print 'CsvExporter.do',
        obj = self.parent
        attrsman = self.get_attrsman()
        sep = self.sep
        #logger = self._logger

        obj.export_csv(self.csvfilepath, sep=self.sep,
                       #ids = None,
                       attrconfigs=None,
                       groupnames=None,
                       show_parentesis=self.show_parentesis,
                       is_export_not_save=True,  # export attr, also non save
                       # name_id='ID',
                       is_header=self.is_header)
